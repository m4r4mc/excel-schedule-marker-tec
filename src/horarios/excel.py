"""Genera el archivo Excel"""

from itertools import product
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from .constants import DIAS_SEMANA, COLORES
from .combinaciones import choque


def borde_f():
    s = Side(style="thin", color="999999")
    return Border(left=s, right=s, top=s, bottom=s)


def slots_30min():
    """Genera celdas de 30 min desde 7:00 hasta 22:00"""
    slots = []
    total = 7 * 60   # 7:00
    fin = 22 * 60    # 22:00
    while total < fin:
        slots.append(total)
        total += 30
    return slots


def min_a_label(m):
    h, mi = divmod(m, 60)
    return f"{h}:{mi:02d}"


def generar_excel(cursos, archivo_salida="horarios_finales.xlsx"):
    wb = Workbook()
    del wb["Sheet"]

    combos = list(product(*[c["grupos"] for c in cursos]))
    combos = [c for c in combos if not hay_choque(c)]
    combos.sort(key=lambda c: sum(g["interes"] for g in c), reverse=True)

    ws_resumen = wb.create_sheet("Resumen")
    ws_resumen.append(["Ranking", "Score"])
    for i, c in enumerate(combos):
        ws_resumen.append([i + 1, sum(g["interes"] for g in c)])

    SLOTS = slots_30min()

    for i, combo in enumerate(combos, 1):
        ws = wb.create_sheet(f"Horario {i}"[:31])

        ws.merge_cells("A1:G1")
        ws["A1"] = f"HORARIO OPTIMO #{i}  |  Score: {sum(g['interes'] for g in combo)}"
        ws["A1"].font = Font(bold=True, size=14, color="FFFFFF")
        ws["A1"].fill = PatternFill("solid", start_color="2C3E50")
        ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 24

        ws.cell(row=2, column=1, value="Hora").font = Font(bold=True, color="FFFFFF")
        ws.cell(row=2, column=1).fill = PatternFill("solid", start_color="1E3A5F")
        ws.cell(row=2, column=1).border = borde_fino()
        for ci, dia in enumerate(DIAS_SEMANA, 2):
            c = ws.cell(row=2, column=ci, value=dia)
            c.font = Font(bold=True, color="FFFFFF")
            c.fill = PatternFill("solid", start_color="1E3A5F")
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.border = borde_f()

        ws.column_dimensions["A"].width = 14
        for ci in range(2, 2 + len(DIAS_SEMANA)):
            ws.column_dimensions[ws.cell(row=2, column=ci).column_letter].width = 20

        for ri, slot in enumerate(SLOTS, 3):
            label = f"{min_a_label(slot)} - {min_a_label(slot + 30)}"
            c = ws.cell(row=ri, column=1, value=label)
            c.font = Font(size=9)
            c.alignment = Alignment(horizontal="right", vertical="center")
            c.border = borde_fino()
            for ci in range(2, 2 + len(DIAS_SEMANA)):
                ws.cell(row=ri, column=ci).border = borde_f()

        for gi, g in enumerate(combo):
            color = PatternFill("solid", start_color=COLORES[gi % len(COLORES)])
            curso = next(c for c in cursos if g in c["grupos"])
            texto = f"{curso['nombre']}\nG{g['num']}"
            ini_m, fin_m = g["horario"]["ini_m"], g["horario"]["fin_m"]

            for d in g["dias"]:
                col = DIAS_SEMANA.index(d) + 2
                for ri, slot in enumerate(SLOTS, 3):
                    if ini_m <= slot < fin_m:
                        cell = ws.cell(row=ri, column=col)
                        cell.value = texto
                        cell.fill = color
                        cell.font = Font(color="FFFFFF", bold=True, size=9)
                        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                        cell.border = borde_fino()

        dr = len(SLOTS) + 5
        ws.cell(row=dr, column=1, value="GRUPOS SELECCIONADOS").font = Font(bold=True, color="FFFFFF")
        ws.cell(row=dr, column=1).fill = PatternFill("solid", start_color="1E3A5F")
        for col, h in enumerate(["Curso", "Grupo", "Días", "Horario", "Interés"], 1):
            c = ws.cell(row=dr + 1, column=col, value=h)
            c.font = Font(bold=True, color="FFFFFF")
            c.fill = PatternFill("solid", start_color="2980B9")
            c.border = borde_f()

        for gi, g in enumerate(combo):
            curso = next(c for c in cursos if g in c["grupos"])
            color = PatternFill("solid", start_color=COLORES[gi % len(COLORES)])
            fila = dr + 2 + gi
            valores = [
                curso["nombre"],
                f"Grupo {g['num']}",
                " & ".join(g["dias"]),
                f"{g['horario']['ini']}-{g['horario']['fin']}",
                f"{g['interes']}/10",
            ]
            for col, val in enumerate(valores, 1):
                c = ws.cell(row=fila, column=col, value=val)
                c.fill = color
                c.font = Font(color="FFFFFF", size=10)
                c.border = borde_fino()

        ws.freeze_panes = "B3"

    wb.save(archivo_salida)
    print(f"\n Excel generado correctamente: {archivo_salida}")
